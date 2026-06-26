import openpyxl
from openpyxl.styles import Font, numbers
from process_student import ParseStudent

def dynamic_write(filepath, data: ParseStudent) -> None:
    def safe_text(value: str | None, default: str = "N/A") -> str:
        return value.strip() if value and value.strip() else default

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    bold_text = Font(b=True)
    normal_text = Font(b=False)
    row = 1

    while ws.cell(row=row, column=1).value is not None:
        row += 1

    # Student name cell with proper contact person logic + formatting
    ws.cell(row=row, column=1).value = safe_text(data.student_name)
    ws.cell(row=row, column=1).font = bold_text if data.is_contact_person else normal_text

    # TODO: Fix this so the group numbers work: ws.cell(row=row, column=2).value = row - 5

    # Student requested properties cell
    ws.cell(row=row, column=3).value = safe_text(data.property_one)
    ws.cell(row=row, column=4).value = safe_text(data.property_two)
    ws.cell(row=row, column=5).value = safe_text(data.property_three)

    # Student current address cell with formatting
    cell_current_address = ws.cell(row=row, column=6)
    cell_current_address.value = safe_text(data.current_address)
    cell_current_address.number_format = numbers.FORMAT_TEXT

    # Student phone cell with formatting
    cell_phone = ws.cell(row=row, column=7)
    cell_phone.value = safe_text(str(data.student_phone) if data.student_phone is not None else None)
    cell_phone.number_format = numbers.FORMAT_TEXT

    # Student email cell with formatting
    cell_email = ws.cell(row=row, column=8)
    cell_email.value = safe_text(data.student_email)
    cell_email.number_format = numbers.FORMAT_TEXT

    # Student gpa cell with formatting
    cell_gpa = ws.cell(row=row, column=9)
    try:
        cell_gpa.value = float(data.gpa) if data.gpa and str(data.gpa).strip() else "N/A"
    except (ValueError, TypeError):
        cell_gpa.value = "N/A"
    cell_gpa.number_format = '0.00'

    # Student greek life cell
    ws.cell(row=row, column=10).value = safe_text(data.greek_life)

    # Save the workbook changes
    wb.save(filepath)

    